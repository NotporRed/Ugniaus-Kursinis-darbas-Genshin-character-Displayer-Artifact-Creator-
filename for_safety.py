import openpyxl
import pygame
import sys 
import random
import numpy as np
import json
import os.path
from abc import ABC, abstractmethod

class Game:
    def __init__(self):
        pygame.init()
        self.screen = pygame.display.set_mode((1280, 720))
        self.clock = pygame.time.Clock()
        
        self.game_state_manager = Game_state_manager('selection')
        self.selection = Selection(self.screen, self.game_state_manager)
        self.selection_weapon = Selection_weapon(self.screen, self.game_state_manager)
        self.character_screen = Character_screen(self.screen, self.game_state_manager)
        
        self.states = {'selection':self.selection, 'character_screen':self.character_screen, 'selection_weapon':self.selection_weapon}
    
    def run(self):
        while True:
            event_list = pygame.event.get()
            for event in event_list:
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()
                    
            self.states[self.game_state_manager.get_state()].run(event_list)
                
            pygame.display.update()
            self.clock.tick(144)       
            
class Selection():
    def __init__(self, display, game_state_manager):
        self.display = display
        self.game_state_manager = game_state_manager
        self.user_text = ""
        self.character_list = showcase_caracter.character_list
        self.filterred_list = self.character_list
        self.button = []
        self.index = 0
  
    def fillter_the_list(self):
        self.filterred_list = [x for x in self.character_list if x.startswith(self.user_text)]
        if not len(self.filterred_list):
            self.filterred_list.append(self.character_list[self.index])
        
        if self.index > len(self.filterred_list):
            self.index = len(self.filterred_list)-1
            
        
    def create_button(self):
        self.button.append(Button(self.display, (104,71,104), 415, 500, 450, 150, self.filterred_list[self.index]))
        self.button.append(Button(self.display, (104,71,104), 415, 450, 450, 50, self.user_text))     
        self.button.append(Button(self.display, (104,71,104), 255, 500, 150, 150, '<'))
        self.button.append(Button(self.display, (104,71,104), 875, 500, 150, 150, '>'))
        
    def run(self, event_list):
        self.display.fill((30, 1, 30))
        self.create_button()
        self.button[0].update_text(self.filterred_list[self.index])
        
        png = pygame.image.load("Icons/" + self.filterred_list[self.index].replace(" ", "_") + "_Icon.png")
        self.display.blit(png, (512,194))
        
        for event in event_list:
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_BACKSPACE:
                    self.user_text = self.user_text[:-1]
                    self.fillter_the_list()
                else:
                    if event.key in range( pygame.K_a, pygame.K_z + 1 ) or event.key == pygame.K_SPACE:
                        if len(self.user_text) < 15:
                            self.user_text += event.unicode
                            self.fillter_the_list()
                        
                        else:
                            self.user_text = self.user_text[:-1]
                            self.user_text += event.unicode
                            self.fillter_the_list()
                
            if event.type == pygame.MOUSEBUTTONUP:
                if self.button[0].is_hovered():
                    showcase_caracter.set_current(self.button[0].give_text())
                    self.game_state_manager.set_state('character_screen')
                    
                elif self.button[2].is_hovered():
                    if len(self.filterred_list) > 1 and self.index > 1:
                        self.index -= 1
                    else:
                        self.index = len(self.filterred_list)-1
                    
                elif self.button[3].is_hovered():
                    if self.index < len(self.filterred_list)-1:
                        self.index += 1
                    else:
                        self.index = 0
            
class Selection_weapon():
    def __init__(self, display, game_state_manager):
        self.display = display
        self.game_state_manager = game_state_manager
        self.user_text = ""
        self.weapon_list = item_Catalogue.weapon_list
        self.filterred_list = self.weapon_list
        self.button = []
        self.index = 0
  
    def fillter_the_list(self):
        self.filterred_list = [x for x in self.weapon_list if x.startswith(self.user_text)]
        if not len(self.filterred_list):
            self.filterred_list.append(self.weapon_list[self.index])
            
        if self.index > len(self.filterred_list):
            self.index = len(self.filterred_list)-1
             
    def create_button(self):
        self.button.append(Button(self.display, (104,71,104), 415, 500, 450, 150, self.filterred_list[self.index]))
        self.button.append(Button(self.display, (104,71,104), 415, 450, 450, 50, self.user_text))     
        self.button.append(Button(self.display, (104,71,104), 255, 500, 150, 150, '<'))
        self.button.append(Button(self.display, (104,71,104), 875, 500, 150, 150, '>'))
        
    def run(self, event_list):
        self.display.fill((30, 1, 30))
        self.create_button()
        
        self.button[0].update_text(self.filterred_list[self.index])
        
        for event in event_list:
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_BACKSPACE:
                    self.user_text = self.user_text[:-1]
                    self.fillter_the_list()
                else:
                    if event.key in range( pygame.K_a, pygame.K_z + 1 ) or event.key == pygame.K_SPACE:
                        if len(self.user_text) < 15:
                            self.user_text += event.unicode
                            self.fillter_the_list()
                        
                        else:
                            self.user_text = self.user_text[:-1]
                            self.user_text += event.unicode
                            self.fillter_the_list()
                
            if event.type == pygame.MOUSEBUTTONUP:
                if self.button[0].is_hovered():
                    item_Catalogue.set_current(self.button[0].give_text())
                    self.game_state_manager.set_state('character_screen')
                    
                elif self.button[2].is_hovered():
                    if len(self.filterred_list) > 1 and self.index > 1:
                        self.index -= 1
                    else:
                        self.index = len(self.filterred_list)-1
                    
                elif self.button[3].is_hovered():
                    if self.index < len(self.filterred_list)-1:
                        self.index += 1
                    else:
                        self.index = 0
            
            if event.type == pygame.K_DELETE:
                self.game_state_manager.set_state('character_screen')
                               
class Character_screen():
    def __init__(self, display, game_state_manager):
        self.character = showcase_caracter.main_character
        self.weapon = item_Catalogue.current_weapon
        self.scaling = multiplier
        self.display = display
        self.game_state_manager = game_state_manager
        self.level = 1
        self.asc = 1
        self.button = []
        self.factory = Artifact_Factory()
        self.artifact_1 = None
        self.artifact_2 = None
        self.artifact_3 = None
        self.artifact_4 = None
        self.artifact_5 = None
        
    def text_preset(self, atr, atr_mult):
        if atr == "HP%":
            return self.print_text(atr, hp = atr_mult)
        elif atr == "ATK%":
            return self.print_text(atr, atk = atr_mult)
        elif atr == "DEF%":
            return self.print_text(atr, defence = atr_mult)
        elif atr == "CR":
            return self.print_text(atr, cr = atr_mult)
        elif atr == "CD":
            return self.print_text(atr, cd = atr_mult)
        elif atr == "EM":
            return self.print_text(atr, m = atr_mult)
        elif atr == "ER":
            return self.print_text(atr, er = atr_mult)
        elif atr == "Healing_Bonus":   
            return self.print_text(healing_bonus = atr_mult)
        elif atr == "Phys DMG%":
            return self.print_text(atr, phys = atr_mult)
        else:
            return self.print_text(atr, dmg = atr_mult)
        
    def print_text(self, atr, hp=0, atk=0, defence=0, cr=0, cd=0, em=0, er=0, healing_bonus=0, phys=0 , dmg=0):
        lv_mult = float(self.scaling.level_multiplier(self.level, self.character.give_item(3)))
        lv_asc = float(self.scaling.level_ascention_multiplier(self.asc))
        
        Text(self.display, str(self.character.give_item(1)), ('white'), 40, 30, 40)
        Text(self.display, "Level: " + str(self.level) + "/90", ('white'), 40, 80, 35)
        Text(self.display, "A" + str(self.asc), ('white'), 40, 130, 20)
        
        text_hp = self.character.give_item(5) * lv_mult + lv_asc * self.scaling.max_asc_value(self.character.give_item(1), hp=1) + hp + (self.weapon.give_value() * 100 if self.weapon.give_sub() == "HP%" else 0)    
        text_atk = self.character.give_item(6) * lv_mult + lv_asc * self.scaling.max_asc_value(self.character.give_item(1), atk=1) + atk + self.weapon.give_atk() + (self.weapon.give_value() * 100 if self.weapon.give_sub() == "ATK%" else 0)
        text_defence = self.character.give_item(7) * lv_mult + lv_asc * self.scaling.max_asc_value(self.character.give_item(1), defence=1) + defence + (self.weapon.give_value() * 100 if self.weapon.give_sub() == "DEF%" else 0)
        text_crit_rate = (0.05 + cr) * 100 + (self.weapon.give_value() * 100 if self.weapon.give_sub() == "CR" else 0)
        text_crit_dmg = (0.5 + cd) * 100 + (self.weapon.give_value() * 100 if self.weapon.give_sub() == "CD" else 0)
        text_elemental_mastery = 0 + em + (self.weapon.give_value() if self.weapon.give_sub() == "EM" else 0)
        text_energy_recharge = (1 + er) * 100 + (self.weapon.give_value() * 100 if self.weapon.give_sub() == "ER" else 0)
        text_healing_bonus = (0 + healing_bonus) * 100
        text_dmg = (dmg)*100
        
        all_artifacts = []
        if self.artifact_1:
            all_artifacts.append(self.artifact_1.give_artifact())
        if self.artifact_2:
            all_artifacts.append(self.artifact_2.give_artifact())
        if self.artifact_3:
            all_artifacts.append(self.artifact_3.give_artifact())
        if self.artifact_4:
            all_artifacts.append(self.artifact_4.give_artifact())
        if self.artifact_5:
            all_artifacts.append(self.artifact_5.give_artifact())
        
        if all_artifacts is not None:
            for x in range(len(all_artifacts)):
                for key in all_artifacts[x].keys():
                    match key:
                        case "HP%":
                            text_hp += (all_artifacts[x].get("HP%"))
                        
                        case "HP":
                            text_hp += all_artifacts[x].get("HP")
                            
                        case "ATK%":
                            text_atk += (all_artifacts[x].get("ATK%"))

                        case "ATK":
                            text_atk += all_artifacts[x].get("ATK")
                        
                        case "DEF%":
                            text_defence += (all_artifacts[x].get("DEF%"))
                            
                        case "DEF":
                            text_defence += all_artifacts[x].get("DEF")
                        
                        case "ER":
                            text_energy_recharge += (all_artifacts[x].get("ER"))
                            
                        case "EM":
                            text_elemental_mastery += all_artifacts[x].get("EM")
                        
                        case "CR":
                            text_crit_rate += (all_artifacts[x].get("CR"))

                        case "CD":
                            text_crit_dmg += (all_artifacts[x].get("CD"))
                            
                        case "DMG%":
                            text_dmg += (all_artifacts[x].get("DMG%"))
                            atrifact_dmg = all_artifacts[x].get("DMG%")
        
        Text(self.display, "Hp: " + str(round(text_hp)), 'white', 40, 170, 25)
    
        Text(self.display, "Atk: " + str(round(text_atk)), ('white'), 40, 200, 25)
        Text(self.display, "Def: " + str(round(text_defence)), ('white'), 40, 230, 25)
        
        Text(self.display, "Crit rate: " + str(round(text_crit_rate, 1)) + "%", ('white'), 40, 270, 25)
        Text(self.display, "Crit dmg: " + str(round(text_crit_dmg, 1)) + "%", ('white'), 40, 300, 25)
        
        Text(self.display, "Elemental mastery: " + str(round(text_elemental_mastery)), ('white'), 40, 340, 25)
        Text(self.display, "Energy recharge: " + str(round(text_energy_recharge,1)) + "%", ('white'), 40, 370, 25)
        Text(self.display, "Healing bonus: " + str(round(text_healing_bonus, 1)) + "%", ('white'), 40, 400, 25)
        
        if phys != 0:
            Text(str(self.display, self.character.give_item(2)) + "DMG %: +" + str(round((dmg+atrifact_dmg)*100,1)), ('white'), 40, 430, 25)
        else:
            Text(self.display, self.character.give_item(2) + ": +" + str(round((dmg)*100 + self.weapon.give_value*100 if self.weapon.give_sub == "Phys DMG%" else 0, 1)), ('white'), 40, 430, 25)
        
        if em == 0:
            Text(self.display, "Ascenion Stat: +" + str(round(hp + atk + defence + cr + cd + er + healing_bonus + dmg,1)) + " " + atr, ('white'), 40, 460, 25)
        else:
            Text(self.display, "Ascenion Stat: " + str(round(em)) + " " + atr, ('white'), 50, 450, 25)            
              
    def create_button(self):
        self.button = []
        self.button.append(Button(self.display, (230,230,236), 50, 625, 75, 50, '<<'))
        self.button.append(Button(self.display, (230,230,236), 135, 625, 75, 50, '<'))
        self.button.append(Button(self.display, (230,230,236), 220, 625, 75, 50, '>'))
        self.button.append(Button(self.display, (230,230,236), 305, 625, 75, 50, '>>'))
        self.button.append(Button(self.display, (230,230,236), 135, 550, 75, 50, '<'))
        self.button.append(Button(self.display, (230,230,236), 220, 550, 75, 50, '>'))
        self.button.append(Button(self.display, (230,230,236), 475, 650, 330, 50, 'Change Weapon'))
        self.button.append(Button(self.display, (230,230,236), 475, 300, 330, 330, self.weapon.give_name()))
        self.button.append(Button(self.display, (230,230,236), 50, 550, 75, 50, 'min'))
        self.button.append(Button(self.display, (230,230,236), 305, 550, 75, 50, 'max'))
        self.button.append(Button(self.display, (230,230,236), 825, 650, 75, 50, '1'))
        self.button.append(Button(self.display, (230,230,236), 905, 650, 75, 50, '2'))
        self.button.append(Button(self.display, (230,230,236), 985, 650, 75, 50, '3'))
        self.button.append(Button(self.display, (230,230,236), 1065, 650, 75, 50, '4'))
        self.button.append(Button(self.display, (230,230,236), 1145, 650, 75, 50, '5'))
        self.button.append(Button(self.display, (230,230,236), 1145, 550, 75, 50, 'save'))
        self.button.append(Button(self.display, (230,230,236), 1145, 450, 75, 50, 'load'))
        
    def run(self, event_list):
        self.display.fill((34,28,105))     
        self.character = showcase_caracter.main_character
        self.weapon = item_Catalogue.current_weapon
        
        pygame.draw.rect(self.display, (106,107,141), (20, 20, 450, 680), 0)
        pygame.draw.rect(self.display, (8,9,66), (20, 20, 450, 680), 5)
        
        atr_mult = self.scaling.atribute_position(self.character.give_item(4), self.asc, self.character.give_item(3))
        self.text_preset(self.character.give_item(4), atr_mult)
        
        self.create_button()
        self.button[7].update_text(self.weapon.give_name())
        
        for event in event_list:
            if event.type == pygame.MOUSEBUTTONUP:
                if self.button[0].is_hovered():
                    item_Catalogue.set_current(self.button[0].give_text())
                    self.game_state_manager.set_state('character_screen')
                    
                elif self.button[2].is_hovered():
                    if len(self.filterred_list) > 1 and self.index > 1:
                        self.index -= 1
                    else:
                        self.index = len(self.filterred_list)-1
                    
                elif self.button[3].is_hovered():
                    if self.index < len(self.filterred_list)-1:
                        self.index += 1
                    else:
                        self.index = 0
                        
            if event.type == pygame.MOUSEBUTTONUP:
                if self.button[0].is_hovered():
                    if self.level > 11:
                        self.level -= 10
                    else:
                        self.level = 1
                        
                if self.button[1].is_hovered():
                    if self.level > 1:
                        self.level -= 1
                
                if self.button[2].is_hovered():
                    if self.level < 90:
                        self.level += 1
                            
                if self.button[3].is_hovered():
                    if self.level < 80:
                        self.level += 10
                    else:
                        self.level = 90
                        self.update = True
                        
                if self.button[4].is_hovered():
                    if self.asc > 1:
                        self.asc -= 1
                        
                if self.button[5].is_hovered():
                    if self.asc < 6:
                        self.asc += 1
                        self.key_clicked = True
                        self.update = True
                        
                if self.button[6].is_hovered():
                    self.game_state_manager.set_state('selection_weapon')    
                    
                if self.button[8].is_hovered():
                    self.asc = 1
                    self.level = 1

                if self.button[9].is_hovered():
                    self.asc = 6
                    self.level = 90
                
                if self.button[10].is_hovered():
                    self.artifact_1 = self.factory.create_artifact(1)
                        
                if self.button[11].is_hovered():
                    self.artifact_2 = self.factory.create_artifact(2)
                    
                if self.button[12].is_hovered():
                    self.artifact_3 = self.factory.create_artifact(3)
                    
                if self.button[13].is_hovered():
                    self.artifact_4 = self.factory.create_artifact(4)
                    
                if self.button[14].is_hovered():
                    self.artifact_5 = self.factory.create_artifact(5)
                    
                if self.button[15].is_hovered():
                    for artifact in [self.artifact_1, self.artifact_2, self.artifact_3, self.artifact_4, self.artifact_5]:
                        if artifact:
                            artifact.save_to_file()
                    return

                if self.button[16].is_hovered():
                    loaded_artifacts = Artifact.load_from_file()
                    if len(loaded_artifacts) >= 5:
                        self.artifact_1 = loaded_artifacts[0]
                        self.artifact_2 = loaded_artifacts[1]
                        self.artifact_3 = loaded_artifacts[2]
                        self.artifact_4 = loaded_artifacts[3]
                        self.artifact_5 = loaded_artifacts[4]
                    else:
                        print("Error: Not enough artifacts loaded from file.")
                                
        key = pygame.key.get_pressed()
        if key[pygame.K_BACKSPACE]:
            self.game_state_manager.set_state('selection')      

class Game_state_manager:
    def __init__(self, current_state):
        self.current_state = current_state
        
    def get_state(self):
        return self.current_state
    
    def set_state(self, state):
       self.current_state = state 

class Text():
    def __init__(self, display, text, color, x, y, size):
        self.text = text.replace("_", " ")
        self.color = color
        self.x = x
        self.y = y
        self.font = pygame.font.Font('Fonts/ja-jp.ttf',size)
        self.show_text(display)
    
    def show_text(self, display):
        block_text = self.font.render(self.text, 1, (self.color))
        display.blit(block_text, (self.x, self.y))

class Button:
    def __init__(self, display, color, x, y, width, height, text):
        self.color = color
        self.x = x
        self.y = y
        self.width = width
        self.height = height
        self.text = text
        self.font = pygame.font.Font('Fonts/ja-jp.ttf',25)
        self.rect = pygame.Rect(x, y, width, height)
        self.draw(display)
        self.display_text(display)
        
    def draw(self, display):
        if  self.is_hovered():
            pygame.draw.rect(display, (tuple(x-50 for x in self.color)), (self.x, self.y, self.width, self.height), 0)
        else:
            pygame.draw.rect(display, self.color, (self.x, self.y, self.width, self.height), 0)
            
        pygame.draw.rect(display, (10, 10, 10), (self.x, self.y, self.width, self.height), 5)

        
    def display_text(self, display):
        block_text = self.font.render(self.text.replace("_", " "), 1, ('black'))
        text_rect = block_text.get_rect(center = (self.x+self.width/2, self.y+self.height/2))
        display.blit(block_text, text_rect)
        
    def is_hovered(self):
        return pygame.MOUSEBUTTONUP and self.rect.collidepoint(pygame.mouse.get_pos())
    
    def give_text(self):
        return self.text
    
    def update_text(self, text):
        self.text = text
         
class Character_stats():
    def __init__(self, name, element, rarty, hp, atk, defence, ascention_stat):
        self.name = name
        self.element = element
        self.rarty = rarty
        self.ascention_stat = ascention_stat
        
        self.hp = hp
        self.atk = atk
        self.defence = defence
    
    def give_item(self, item):
        if item == 1:
            return self.name
        
        if item == 2:
            return self.element
        
        if item == 3:
            return self.rarty
        
        if item == 4:
            return self.ascention_stat
        
        if item == 5:
            return float(self.hp)
        
        if item == 6:
            return float(self.atk)
        
        if item == 7:
            return float(self.defence)
            
class Weapon():
    def __init__(self, name, atk, substat, value) -> None:
        self.name = name
        self.atk = atk
        self.substat = substat
        self.value = value
        
    def give_name(self):
        return self.name
    
    def give_atk(self):
        return self.atk
    
    def give_sub(self):
        return self.substat

    def give_value(self):
        return self.value

class Artifact:
    def __init__(self, slot):
        self.slot = slot
        self.subs = {}
        self.generate_subs()
        self.enhance()
        
    def generate_subs(self):
        if self.slot == 1:
            self.subs["HP"] = 4780
            
        if self.slot == 2:
            self.subs["ATK"] = 311
            
        if self.slot == 3:
            self.subs["ATK%"] = 46.6
            
        if self.slot == 4:
            self.subs["DMG%"] = 46.6
            
        if self.slot == 5:
            x = random.randint(1, 2)
            if x == 1:
                self.subs["CR"]= 31.1
            else:
                self.subs["CD"] = 62.2
            
        for n in range(4):
            x = self.random_atr()
            if x and x[0] in self.subs.keys():  # Check if x is not None before accessing keys
                self.subs[x[0]] += x[1]
            else:
                if x:  # Check if x is not None before creating a new key
                    self.subs[x[0]] = x[1]
               
    def random_atr(self):
        x = random.randint(1, 4088)
        if x < 560:
            return "HP", 239
        elif x < 1120:
            return "ATK", 17.51
        elif x < 1680:
            return "DEF", 20.83
        elif x < 2096:
            return "HP%", 5.25
        elif x < 2512:
            return "ATK%", 5.56
        elif x < 2928:
            return "DEF%", 6.93
        elif x < 3344:
            return "ER", 5.83
        elif x < 3760:
            return "EM", 20.93
        elif x < 4088:
            return "CR", 3.5
        elif x < 1120:
            return "CD", 6.99
        
    def enhance(self):
        for n in range(5):
            upgrade_list = list(self.subs)
            upgrade_list.pop(0)
            x = random.choice(upgrade_list)
            self.subs[x] *= 1.5

    def give_artifact(self):
        return self.subs
    
    def save_to_file(self):
        filename = "Artifacts.txt"
        try:
            with open(filename, "a") as fp:
                for stat, value in self.subs.items():
                    fp.write(f"{stat}:{value}\n")
                    fp.write("\n")
        except OSError as e:
            print(f"Error saving artifact to file: {e}")

    @staticmethod
    def load_from_file():
        filename = "Artifacts.txt"
        artifacts = []
        try:
            with open(filename, 'r') as f:
                lines = f.readlines()
                current_artifact = {}
                for line in lines:
                    line = line.strip()
                    if not line:
                        if current_artifact:
                            artifacts.append(Artifact(1))
                            artifacts[-1].subs = current_artifact
                            current_artifact = {}
                    else:
                        stat, value = line.split(':')
                        current_artifact[stat] = float(value)
        except (FileNotFoundError, ValueError) as e:
            print(f"Error loading artifacts from file: {e}")
        return artifacts
    
class Artifact_Factory:
    def create_artifact(self, slot):
        return Artifact(slot)
    
class Check_import():       
    def Check_if_imported(object):
            check = object
            if check:
                print("Files imported")
            else:
                print("Importing files")        
            return(check)

class Loder(ABC):
    @abstractmethod
    def import_to_list(self):
        pass
    
    @abstractmethod
    def set_current(self, name):
        pass
    
    @abstractmethod
    def give_item_list(self):
        pass
    
class Level_multiplier(Loder):
    def __init__(self):
        self.lv_mult_list_4 = []
        self.lv_mult_list_5 = []
        self.lv_mult_asc = []
        
        self.ba_name = []
        self.ba_mult_4 = []
        self.ba_mult_5 = []
        self.ba_mult_asc = []
        
        self.Max_name = []
        self.Max_hp = []
        self.Max_atk = []
        self.Max_def = []
        
        self.imported = False
        self.import_to_list(self.imported) 
      
    @Check_import.Check_if_imported
    def import_to_list(self, check):
        if not check:
            wb = openpyxl.load_workbook('Genshin_character_info.xlsx')
            ws = wb['Level_Scalling']
            
            for row in ws.iter_cols(min_row = 3, max_row = 92, min_col = 2, max_col = 3):
                    for cell in row:
                        if cell.column == 2:
                            self.lv_mult_list_4.append(cell.value)
                        
                        elif cell.column == 3:
                            self.lv_mult_list_5.append(cell.value)
                    
            for row in ws.iter_cols(min_row = 3, max_row = 8, min_col = 6, max_col = 6):
                for cell in row:
                    self.lv_mult_asc.append(cell.value)
                    
            for row in ws.iter_cols(min_row = 3, max_row = 18, min_col = 8, max_col = 10):
                for cell in row:
                    if cell.column == 8:
                        self.ba_name.append(cell.value)
            
                    elif cell.column == 9:
                        self.ba_mult_4.append(cell.value)

                    elif cell.column == 10:
                        self.ba_mult_5.append(cell.value)
                    
            for row in ws.iter_cols(min_row = 3, max_row = 8, min_col = 13, max_col = 13):
                for cell in row:
                    self.ba_mult_asc.append(cell.value)
                    
            for row in ws.iter_cols(min_row = 3, max_row = 87, min_col = 15, max_col = 18):
                for cell in row:
                    if cell.column == 15:
                        self.Max_name.append(cell.value)
                        
                    elif cell.column == 16:
                        self.Max_hp.append(cell.value)
            
                    elif cell.column == 17:
                        self.Max_atk.append(cell.value)

                    elif cell.column == 18:
                        self.Max_def.append(cell.value)   
                                           
        self.imported = True
        
    def set_current(self, name):
        pass
    
    def give_item_list():
        pass
        
    def level_multiplier(self, level, rarity):
        if rarity == "4*":
            return self.lv_mult_list_4[level-1]
        
        else:
            return self.lv_mult_list_5[level-1]
        
    def level_ascention_multiplier(self, asc):
        return self.lv_mult_asc[asc-1]
    
    def bonus_atribute_multiplier(self, atribute, rarity):
        x = range(16)
        for n in x:
            if atribute == self.ba_name[n]:
                if rarity == "4*":
                    return self.ba_mult_4[n]
                elif rarity == "5*":
                    return self.ba_mult_5[n]
                
    def bonus_atribute_ascention_multiplier(self, asc):
        return self.ba_mult_asc[asc-1]
    
    def atribute_position(self, atribute, asc, rarity):
                m = float(self.bonus_atribute_multiplier(atribute, rarity))
                b = float(self.bonus_atribute_ascention_multiplier(asc))
                return m * b
           
    def max_asc_value(self, name, hp = None, atk = None, defence = None):
        for x in range(86):
            if name == self.Max_name[x]:
                if hp != None:
                    return self.Max_hp[x]
                elif atk != None:
                    return self.Max_atk[x]
                elif defence != None:
                    return self.Max_def[x]
  
class Showcase_caracter(Loder):
    def __init__(self):
        self.imported = False
        self.character_catalogue = []
        self.character_list = []
        self.import_to_list(self.imported)
        self.give_item_list()
        self.main_character = self.character_catalogue[0]
            
    @Check_import.Check_if_imported
    def import_to_list(self, check):
        if not check:
            wb = openpyxl.load_workbook('Genshin_character_info.xlsx')
            ws = wb['CHARACTER']
            
            base_stats_data = []
            
            for row in ws.iter_rows(min_row = 2, max_row = 86, max_col = 7):
                for cell in row:
                    base_stats_data.append(cell.value)
                    
                if len(base_stats_data) == 7:
                    self.character_catalogue.append(Character_stats(*base_stats_data))
                    base_stats_data = []   
                                
            wb.close()
                
        self.imported = True

    def set_current(self, name):
        for x in range(85):
            if name == self.character_catalogue[x].give_item(1):
                self.main_character = self.character_catalogue[x]
    
    def give_item_list(self):
        self.character_list = []
        for x in range(85):
             self.character_list.append(self.character_catalogue[x].give_item(1))
        return  self.character_list
                                
class Item_Catalogue(Loder):
    def __init__(self):
        self.imported = False
        self.weapon_catalogue = []
        self.weapon_list = []
        self.import_to_list(self.imported)
        self.give_item_list()
        self.current_weapon = self.weapon_catalogue[0]
    
    @Check_import.Check_if_imported
    def import_to_list(self, check):
        if not check:
            wb = openpyxl.load_workbook('Genshin_character_info.xlsx')
            ws = wb['GS']
            
            weapon_data = []
            artifact_data = []
            
            for row in ws.iter_rows(min_row = 2, max_row = 175, max_col = 4):
                    for cell in row:
                        if cell.value is not None:
                            weapon_data.append(cell.value)

                        if len(weapon_data) == 4:
                            self.weapon_catalogue.append(Weapon(*weapon_data))
                            weapon_data = []
                            
        self.check_if_imported = True
        
    def set_current(self, name):
        for x in range(1, 173):          
            if name == self.weapon_catalogue[x].give_name():
                self.current_weapon = self.weapon_catalogue[x]
    
    def give_item_list(self):
        self.weapon_list = []
        for x in range(1, 173):
             self.weapon_list.append(self.weapon_catalogue[x].give_name())
        return  self.weapon_list
    
    
if __name__ == '__main__':
    multiplier = Level_multiplier()
    showcase_caracter = Showcase_caracter()
    item_Catalogue = Item_Catalogue()
    game = Game()
    game.run()
    